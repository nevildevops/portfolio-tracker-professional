import sys
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.chart import LineChart, BarChart, Reference

FILENAME = "Portfolio_Tracker_Professional.xlsx"

def ensure_data(wb):
    # Ensure Data sheet
    if "Data" in wb.sheetnames:
        ws_data = wb["Data"]
    else:
        ws_data = wb.create_sheet("Data")
        ws_data.append(["Date", "Portfolio Value"])

    # Fill minimal data if missing
    if ws_data.max_row < 3:
        start_date = datetime(2024, 1, 1)
        value = 10000
        for i in range(12):
            value = round(value * (1 + 0.01 * ((i % 5) - 2)), 2)  # simple oscillation
            ws_data.append([start_date + timedelta(days=30*i), value])

    # Ensure Returns sheet
    if "Returns" in wb.sheetnames:
        ws_ret = wb["Returns"]
    else:
        ws_ret = wb.create_sheet("Returns")
        ws_ret.append(["Asset", "Average Return (%)"])

    if ws_ret.max_row < 3:
        for a, r in [("Stocks", 12.4), ("Bonds", 4.1), ("Gold", 6.3), ("Crypto", 21.9)]:
            ws_ret.append([a, r])

    return ws_data, ws_ret

def rebuild_charts(wb, ws_data, ws_ret):
    # Ensure Charts sheet
    if "Charts" in wb.sheetnames:
        ws_charts = wb["Charts"]
        # remove existing charts
        ws_charts._charts = []
    else:
        ws_charts = wb.create_sheet("Charts")

    # Line chart
    line = LineChart()
    line.title = "Cumulative Portfolio Growth"
    line.y_axis.title = "Value"
    line.x_axis.title = "Date"
    data_ref = Reference(ws_data, min_col=2, min_row=1, max_row=ws_data.max_row, max_col=2)
    cats_ref = Reference(ws_data, min_col=1, min_row=2, max_row=ws_data.max_row)
    line.add_data(data_ref, titles_from_data=True)
    line.set_categories(cats_ref)
    line.height = 12
    line.width = 24
    ws_charts.add_chart(line, "A2")

    # Bar chart
    bar = BarChart()
    bar.title = "Average Asset Returns"
    bar.y_axis.title = "Return (%)"
    bar.x_axis.title = "Asset"
    bar_data_ref = Reference(ws_ret, min_col=2, min_row=1, max_row=ws_ret.max_row, max_col=2)
    bar_cats_ref = Reference(ws_ret, min_col=1, min_row=2, max_row=ws_ret.max_row)
    bar.add_data(bar_data_ref, titles_from_data=True)
    bar.set_categories(bar_cats_ref)
    bar.height = 12
    bar.width = 24
    ws_charts.add_chart(bar, "A22")

def main():
    try:
        wb = load_workbook(FILENAME)
    except FileNotFoundError:
        wb = Workbook()
        # default sheet rename
        ws = wb.active
        ws.title = "Data"
        ws.append(["Date", "Portfolio Value"])

    ws_data, ws_ret = ensure_data(wb)
    rebuild_charts(wb, ws_data, ws_ret)
    wb.save(FILENAME)
    print(f"Updated charts in {FILENAME}")

if __name__ == "__main__":
    main()
